from __future__ import annotations

import argparse
import base64
import binascii
import json
import os
import re
import subprocess
import sys
import webbrowser
from copy import copy
from datetime import date, datetime, timedelta
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from threading import Lock
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.datetime import from_excel


ROOT = Path(__file__).resolve().parents[1]
WEB_ROOT = ROOT / "web"
STATE_PATH = Path(os.environ.get("LFP_MONITOR_STATE_PATH", ROOT / "runtime" / "collection-monitor.json"))
MONITOR_PATH = ROOT / "scripts" / "monitor_collections.py"
PURCHASE_TEMPLATE_PATH = WEB_ROOT / "templates" / "purchase-request-template.xlsx"
DESKTOP_PATH = Path.home() / "Desktop"
DELIVERY_STATE_PATH = ROOT / "runtime" / "delivery-management.json"
DELIVERY_CONFIRMATIONS_PATH = WEB_ROOT / "data" / "lidding-delivery-confirmations.xlsx"
PURCHASE_DATA_PATH = WEB_ROOT / "data" / "lidding-purchase-inbound.json"
DELIVERY_LOCK = Lock()


def delivery_key(item_code: str, spec: str) -> str:
    return f"{item_code.strip().upper()}|{spec.strip().upper()}"


def clean_header(value) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).lower()


def column_for(headers: dict, *aliases: str):
    for alias in aliases:
        column = headers.get(clean_header(alias))
        if column:
            return column
    return None


def workbook_year(filename: str) -> int:
    match = re.search(r"(20\d{2})", str(filename or ""))
    if match:
        return int(match.group(1))
    match = re.search(r"(?<!\d)(\d{2})년", str(filename or ""))
    return 2000 + int(match.group(1)) if match else date.today().year


def excel_date_text(value, year_hint: int) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            if isinstance(converted, datetime):
                converted = converted.date()
            if isinstance(converted, date):
                return converted.isoformat()
        except (TypeError, ValueError, OverflowError):
            pass
    raw = str(value).strip()
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(raw, pattern).date().isoformat()
        except ValueError:
            pass
    match = re.fullmatch(r"(\d{1,2})[./-](\d{1,2})", raw)
    if match:
        try:
            return date(year_hint, int(match.group(1)), int(match.group(2))).isoformat()
        except ValueError:
            return ""
    return ""


def split_identifiers(value) -> list[str]:
    return [part.strip() for part in re.split(r"[\r\n,;]+", str(value or "")) if part.strip()]


def read_purchase_items() -> list[dict]:
    try:
        payload = json.loads(PURCHASE_DATA_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError, TypeError):
        return []
    return payload.get("items", []) if isinstance(payload, dict) else []


def purchase_item_key(item: dict) -> str:
    return delivery_key(str(item.get("itemCode") or ""), str(item.get("specification") or ""))


def purchase_dates(item: dict) -> set[str]:
    values = {
        item.get("nextDeliveryDate"),
        item.get("nextRequestedDeliveryDate"),
        item.get("nextOrderDeliveryDate"),
    }
    values.update(row.get("requestedDeliveryDate") for row in item.get("requests", []) or [])
    values.update(row.get("deliveryDate") for row in item.get("purchaseOrders", []) or [])
    return {str(value)[:10] for value in values if value}


def find_workbook_header(workbook, required_groups):
    for sheet in workbook.worksheets:
        for row_number in range(1, min(sheet.max_row, 40) + 1):
            headers = {
                clean_header(cell.value): cell.column
                for cell in sheet[row_number]
                if cell.value not in (None, "")
            }
            if all(any(clean_header(alias) in headers for alias in aliases) for aliases in required_groups):
                return sheet, row_number, headers
    return None, None, None


def parse_delivery_workbook(content: bytes, filename: str):
    workbook = load_workbook(BytesIO(content), data_only=True)
    purchase_items = read_purchase_items()
    year_hint = workbook_year(filename)
    stats = {"parsed": 0, "blank": 0, "struck": 0, "unmatched": 0, "conflicts": 0}

    sheet, header_row, headers = find_workbook_header(workbook, (("품목코드",), ("납기조정일",)))
    if sheet:
        stats.update({"format": "management", "formatLabel": "납기관리 리스트"})
        columns = {
            "itemCode": column_for(headers, "품목코드"),
            "itemName": column_for(headers, "품목명"),
            "spec": column_for(headers, "규격"),
            "requestedDate": column_for(headers, "납기요청일"),
            "confirmedDate": column_for(headers, "납기확정일"),
            "adjustedDate": column_for(headers, "납기조정일"),
            "note": column_for(headers, "비고"),
            "requestNos": column_for(headers, "구매의뢰번호"),
            "orderNos": column_for(headers, "발주번호"),
            "waitingStatus": column_for(headers, "대기상태"),
        }
        active_request_nos = {
            str(row.get("requestNo") or "").strip()
            for item in purchase_items
            for row in (item.get("requests", []) or []) + (item.get("purchaseOrders", []) or [])
            if row.get("requestNo")
        }
        active_order_nos = {
            str(row.get("purchaseOrderNo") or "").strip()
            for item in purchase_items
            for row in item.get("purchaseOrders", []) or []
            if row.get("purchaseOrderNo")
        }
        rows = []
        for row_number in range(header_row + 1, sheet.max_row + 1):
            item_code = str(sheet.cell(row_number, columns["itemCode"]).value or "").strip()
            if not item_code:
                continue
            adjusted = excel_date_text(sheet.cell(row_number, columns["adjustedDate"]).value, year_hint)
            note = str(sheet.cell(row_number, columns["note"]).value or "").strip() if columns["note"] else ""
            if not adjusted and not note:
                stats["blank"] += 1
                continue
            request_nos = split_identifiers(sheet.cell(row_number, columns["requestNos"]).value) if columns["requestNos"] else []
            order_nos = split_identifiers(sheet.cell(row_number, columns["orderNos"]).value) if columns["orderNos"] else []
            if request_nos and active_request_nos and not any(number in active_request_nos for number in request_nos):
                stats["unmatched"] += 1
                continue
            if order_nos and active_order_nos and not any(number in active_order_nos for number in order_nos):
                stats["unmatched"] += 1
                continue
            rows.append({
                "itemCode": item_code,
                "itemName": str(sheet.cell(row_number, columns["itemName"]).value or "").strip() if columns["itemName"] else "",
                "spec": str(sheet.cell(row_number, columns["spec"]).value or "").strip() if columns["spec"] else "",
                "requestedDate": excel_date_text(sheet.cell(row_number, columns["requestedDate"]).value, year_hint) if columns["requestedDate"] else "",
                "confirmedDate": excel_date_text(sheet.cell(row_number, columns["confirmedDate"]).value, year_hint) if columns["confirmedDate"] else "",
                "adjustedDate": adjusted,
                "note": note,
                "requestNos": request_nos,
                "orderNos": order_nos,
                "waitingStatus": str(sheet.cell(row_number, columns["waitingStatus"]).value or "").strip() if columns["waitingStatus"] else "",
            })
        stats["parsed"] = len(rows)
        return rows, stats

    sheet, header_row, headers = find_workbook_header(
        workbook, (("규격",), ("요청납기", "요청납기일"), ("조정납기", "조정납기일"))
    )
    if not sheet:
        raise ValueError("지원하는 납기관리 또는 납품일정 양식을 찾지 못했습니다.")

    stats.update({"format": "pnp-schedule", "formatLabel": "피앤피 납품일정"})
    code_column = column_for(headers, "규격")
    requested_column = column_for(headers, "요청납기", "요청납기일")
    adjusted_column = column_for(headers, "조정납기", "조정납기일")
    rows = []
    for row_number in range(header_row + 1, sheet.max_row + 1):
        source_cells = [cell for cell in sheet[row_number] if cell.value not in (None, "")]
        if any(bool(cell.font and cell.font.strike) for cell in source_cells):
            stats["struck"] += 1
            continue
        item_code = str(sheet.cell(row_number, code_column).value or "").strip()
        if not item_code:
            continue
        requested = excel_date_text(sheet.cell(row_number, requested_column).value, year_hint)
        adjusted = excel_date_text(sheet.cell(row_number, adjusted_column).value, year_hint)
        if not adjusted:
            stats["blank"] += 1
            continue
        candidates = {
            purchase_item_key(item): item
            for item in purchase_items
            if str(item.get("itemCode") or "").strip().upper() == item_code.upper()
            and (not requested or requested in purchase_dates(item))
        }
        if not candidates and requested:
            candidates = {
                purchase_item_key(item): item
                for item in purchase_items
                if str(item.get("itemCode") or "").strip().upper() == item_code.upper()
            }
        if not candidates:
            stats["unmatched"] += 1
            continue
        if len(candidates) > 1:
            stats["conflicts"] += 1
            continue
        item = next(iter(candidates.values()))
        request_nos = sorted({str(row.get("requestNo") or "").strip() for row in (item.get("requests", []) or []) + (item.get("purchaseOrders", []) or []) if row.get("requestNo")})
        order_nos = sorted({str(row.get("purchaseOrderNo") or "").strip() for row in item.get("purchaseOrders", []) or [] if row.get("purchaseOrderNo")})
        wait_types = []
        if float(item.get("inboundWaitQty") or 0) > 0:
            wait_types.append("입고대기")
        if float(item.get("purchaseWaitQty") or 0) > 0:
            wait_types.append("발주대기")
        rows.append({
            "itemCode": item.get("itemCode", item_code),
            "itemName": item.get("itemName", ""),
            "spec": item.get("specification", ""),
            "requestedDate": requested,
            "confirmedDate": "",
            "adjustedDate": adjusted,
            "note": "",
            "requestNos": request_nos,
            "orderNos": order_nos,
            "waitingStatus": "+".join(wait_types),
        })
    stats["parsed"] = len(rows)
    return rows, stats


def read_delivery_state() -> dict:
    try:
        payload = json.loads(DELIVERY_STATE_PATH.read_text(encoding="utf-8"))
        if isinstance(payload, dict) and isinstance(payload.get("records"), dict):
            return payload
    except (OSError, json.JSONDecodeError):
        pass
    return {"updatedAt": "", "records": {}}


def write_delivery_state(payload: dict) -> None:
    DELIVERY_STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload["updatedAt"] = datetime.now().astimezone().isoformat(timespec="seconds")
    temporary = DELIVERY_STATE_PATH.with_suffix(".tmp")
    temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(DELIVERY_STATE_PATH)


def write_delivery_confirmations(records: dict) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "납기확정일"
    sheet.append(["품목코드", "규격", "납기확정일"])
    for record in sorted(records.values(), key=lambda value: (value.get("itemCode", ""), value.get("spec", ""))):
        confirmed_date = str(record.get("confirmedDate") or "").strip()
        if confirmed_date:
            sheet.append([record.get("itemCode", ""), record.get("spec", ""), confirmed_date])
    DELIVERY_CONFIRMATIONS_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(DELIVERY_CONFIRMATIONS_PATH)


class DashboardHandler(SimpleHTTPRequestHandler):
    def end_headers(self) -> None:
        path = urlparse(self.path).path
        if path.endswith((".js", ".css", ".png", ".jpg", ".jpeg", ".svg", ".woff", ".woff2")):
            self.send_header("Cache-Control", "public, max-age=3600")
        elif not path.startswith("/api/"):
            self.send_header("Cache-Control", "no-cache")
        super().end_headers()

    def send_json(self, status: int, payload: dict) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self) -> None:
        path = urlparse(self.path).path
        if path == "/api/monitor-status":
            try:
                payload = json.loads(STATE_PATH.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                payload = {"running": False, "lastStatus": "not_started"}
            self.send_json(200, payload)
            return
        if path == "/api/delivery-management":
            with DELIVERY_LOCK:
                payload = read_delivery_state()
            self.send_json(200, {
                "ok": True,
                "updatedAt": payload.get("updatedAt", ""),
                "records": list(payload.get("records", {}).values()),
            })
            return
        super().do_GET()

    def do_POST(self) -> None:
        path = urlparse(self.path).path
        supported = {
            "/api/refresh",
            "/api/export-purchase-request",
            "/api/delivery-management/export",
            "/api/delivery-management/update",
            "/api/delivery-management/import",
            "/api/delivery-management/note",
            "/api/delivery-management/reconcile",
        }
        if path not in supported:
            self.send_json(404, {"ok": False, "error": "not_found"})
            return
        try:
            length = int(self.headers.get("Content-Length", "0"))
            body = json.loads(self.rfile.read(length) or b"{}")
        except (ValueError, json.JSONDecodeError):
            self.send_json(400, {"ok": False, "error": "invalid_json"})
            return

        if path == "/api/export-purchase-request":
            self.export_purchase_request(body)
            return
        if path == "/api/delivery-management/export":
            self.export_delivery_management(body)
            return
        if path == "/api/delivery-management/update":
            self.update_delivery_management(body)
            return
        if path == "/api/delivery-management/import":
            self.import_delivery_management(body)
            return
        if path == "/api/delivery-management/note":
            self.append_delivery_note(body)
            return
        if path == "/api/delivery-management/reconcile":
            self.reconcile_delivery_management(body)
            return

        scope = str(body.get("scope") or "all")
        if scope not in {"aps", "inventory", "bom", "all"}:
            self.send_json(400, {"ok": False, "error": "invalid_scope"})
            return
        creation_flags = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
        subprocess.Popen(
            [sys.executable, str(MONITOR_PATH), "--once", scope, "--reason", "dashboard_manual"],
            cwd=ROOT,
            env=os.environ.copy(),
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=creation_flags,
        )
        self.send_json(202, {"ok": True, "accepted": True, "scope": scope})

    def export_purchase_request(self, body: dict) -> None:
        raw_rows = body.get("rows")
        if not isinstance(raw_rows, list) or not raw_rows or len(raw_rows) > 500:
            self.send_json(400, {"ok": False, "error": "invalid_rows"})
            return

        rows = []
        try:
            for raw in raw_rows:
                item_code = str(raw.get("itemCode") or "").strip()
                spec = str(raw.get("spec") or "").strip()
                quantity = int(float(str(raw.get("quantity") or "0").replace(",", "")))
                if not item_code or quantity <= 0:
                    raise ValueError("invalid_purchase_row")
                rows.append({"itemCode": item_code, "spec": spec, "quantity": quantity})
        except (AttributeError, TypeError, ValueError):
            self.send_json(400, {"ok": False, "error": "invalid_purchase_row"})
            return

        if not PURCHASE_TEMPLATE_PATH.exists():
            self.send_json(500, {"ok": False, "error": "purchase_template_missing"})
            return

        try:
            workbook = load_workbook(PURCHASE_TEMPLATE_PATH)
            sheet = workbook.active
            source_row = 2
            max_column = sheet.max_column
            due_date = date.today() + timedelta(days=14)

            for offset, row in enumerate(rows):
                target_row = source_row + offset
                if target_row != source_row:
                    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
                    for column in range(1, max_column + 1):
                        source = sheet.cell(source_row, column)
                        target = sheet.cell(target_row, column)
                        target.value = source.value
                        if source.has_style:
                            target._style = copy(source._style)
                        if source.hyperlink:
                            target._hyperlink = copy(source.hyperlink)
                        if source.comment:
                            target.comment = copy(source.comment)

                sheet.cell(target_row, 5).value = offset + 1
                sheet.cell(target_row, 6).value = row["itemCode"]
                sheet.cell(target_row, 9).value = row["spec"]
                sheet.cell(target_row, 12).value = "국내(원)"
                sheet.cell(target_row, 13).value = row["quantity"]
                sheet.cell(target_row, 13).number_format = "#,##0"
                sheet.cell(target_row, 14).value = None
                sheet.cell(target_row, 15).value = None
                sheet.cell(target_row, 17).value = "KRW"
                sheet.cell(target_row, 19).value = due_date
                sheet.cell(target_row, 19).number_format = "yyyy-mm-dd"

            no_fill = PatternFill(fill_type=None)
            for row_number in range(1, source_row + len(rows)):
                for column in (6, 9, 13, 19):
                    sheet.cell(row_number, column).fill = copy(no_fill)

            DESKTOP_PATH.mkdir(parents=True, exist_ok=True)
            stem = f"{date.today():%Y-%m-%d}_구매의뢰 리스트"
            destination = DESKTOP_PATH / f"{stem}.xlsx"
            suffix = 1
            while destination.exists():
                destination = DESKTOP_PATH / f"{stem} ({suffix}).xlsx"
                suffix += 1
            workbook.save(destination)
        except Exception as error:
            self.send_json(500, {"ok": False, "error": "export_failed", "detail": str(error)})
            return

        self.send_json(200, {
            "ok": True,
            "count": len(rows),
            "dueDate": due_date.isoformat(),
            "filename": destination.name,
            "path": str(destination),
        })

    def import_delivery_management(self, body: dict) -> None:
        try:
            content = base64.b64decode(str(body.get("contentBase64") or ""), validate=True)
            rows, stats = parse_delivery_workbook(content, str(body.get("filename") or "납기관리.xlsx"))
        except (binascii.Error, ValueError, OSError) as error:
            self.send_json(400, {"ok": False, "error": "invalid_delivery_workbook", "detail": str(error)})
            return
        if not rows:
            self.send_json(200, {"ok": True, **stats, "updated": 0, "noted": 0, "unchanged": stats.get("blank", 0)})
            return
        body["rows"] = rows
        body["importStats"] = stats
        self.update_delivery_management(body)

    def export_delivery_management(self, body: dict) -> None:
        rows = body.get("rows")
        if not isinstance(rows, list) or not rows or len(rows) > 1000:
            self.send_json(400, {"ok": False, "error": "invalid_rows"})
            return

        headers = [
            "품목코드", "품목명", "규격", "입고대기", "발주대기",
            "납기요청일", "납기확정일", "납기조정일", "비고",
            "구매의뢰번호", "발주번호", "대기상태",
        ]
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "납기관리"
        sheet.append(["", "", "", "", "", "", "", "기입 필요", "기입 가능", "", "", ""])
        sheet.append(headers)
        header_fill = PatternFill("solid", fgColor="164B7A")
        for cell in sheet[2]:
            cell.fill = copy(header_fill)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        required_fill = PatternFill("solid", fgColor="FCE8E6")
        optional_fill = PatternFill("solid", fgColor="E6F4EA")
        sheet["H1"].fill = copy(required_fill)
        sheet["H1"].font = Font(color="B42318", bold=True)
        sheet["I1"].fill = copy(optional_fill)
        sheet["I1"].font = Font(color="137044", bold=True)
        sheet["H2"].fill = PatternFill("solid", fgColor="A43A34")
        sheet["I2"].fill = PatternFill("solid", fgColor="267158")
        sheet.row_dimensions[1].height = 22
        sheet.row_dimensions[2].height = 24

        purchase_map = {purchase_item_key(item): item for item in read_purchase_items()}
        detail_rows = []
        for raw in rows:
            item_code = str(raw.get("itemCode") or "").strip()
            spec = str(raw.get("spec") or "").strip()
            purchase = purchase_map.get(delivery_key(item_code, spec), {})
            raw_request_numbers = raw.get("requestNos") if isinstance(raw.get("requestNos"), list) else []
            raw_order_numbers = raw.get("orderNos") if isinstance(raw.get("orderNos"), list) else []
            request_numbers = sorted({str(value).strip() for value in raw_request_numbers if str(value).strip()})
            order_numbers = sorted({str(value).strip() for value in raw_order_numbers if str(value).strip()})
            if not request_numbers:
                request_numbers = sorted({str(item.get("requestNo") or "").strip() for item in (purchase.get("requests", []) or []) + (purchase.get("purchaseOrders", []) or []) if item.get("requestNo")})
            if not order_numbers:
                order_numbers = sorted({str(item.get("purchaseOrderNo") or "").strip() for item in purchase.get("purchaseOrders", []) or [] if item.get("purchaseOrderNo")})
            wait_types = []
            if float(raw.get("inboundWaiting") or 0) > 0:
                wait_types.append("입고대기")
            if float(raw.get("purchaseWaiting") or 0) > 0:
                wait_types.append("발주대기")
            sheet.append([
                item_code,
                str(raw.get("itemName") or "").strip(),
                spec,
                int(float(str(raw.get("inboundWaiting") or "0").replace(",", ""))),
                int(float(str(raw.get("purchaseWaiting") or "0").replace(",", ""))),
                str(raw.get("requestedDate") or "").strip(),
                str(raw.get("confirmedDate") or "").strip(),
                "",
                "",
                "\n".join(request_numbers),
                "\n".join(order_numbers),
                "+".join(wait_types),
            ])
            for request in purchase.get("requests", []) or []:
                if float(request.get("purchaseWaitQty") or 0) > 0:
                    detail_rows.append([item_code, spec, "발주대기", request.get("requestNo"), "", "미발주", request.get("purchaseWaitQty"), request.get("requestedDeliveryDate")])
            for order in purchase.get("purchaseOrders", []) or []:
                if float(order.get("inboundWaitQty") or 0) > 0:
                    detail_rows.append([item_code, spec, "입고대기", order.get("requestNo"), order.get("purchaseOrderNo"), order.get("orderStatus"), order.get("inboundWaitQty"), order.get("deliveryDate")])

        for row_number in range(3, sheet.max_row + 1):
            sheet.cell(row_number, 4).number_format = "#,##0"
            sheet.cell(row_number, 5).number_format = "#,##0"
            for column in (6, 7, 8):
                sheet.cell(row_number, column).number_format = "yyyy-mm-dd"
            sheet.cell(row_number, 8).fill = copy(required_fill)
            sheet.cell(row_number, 9).fill = copy(optional_fill)
            sheet.cell(row_number, 9).alignment = Alignment(wrap_text=True, vertical="top")
            for column in (10, 11):
                sheet.cell(row_number, column).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.freeze_panes = "A3"
        sheet.auto_filter.ref = f"A2:L{sheet.max_row}"
        widths = [14, 38, 18, 14, 14, 16, 16, 16, 44, 24, 24, 18]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width

        detail_sheet = workbook.create_sheet("대기번호 상세")
        detail_sheet.append(["품목코드", "규격", "대기구분", "구매의뢰번호", "발주번호", "상태", "연결수량", "납기일"])
        for detail in detail_rows:
            detail_sheet.append(detail)
        for cell in detail_sheet[1]:
            cell.fill = copy(header_fill)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        detail_sheet.freeze_panes = "A2"
        detail_sheet.auto_filter.ref = f"A1:H{max(detail_sheet.max_row, 1)}"
        for index, width in enumerate([14, 18, 14, 22, 22, 16, 16, 16], start=1):
            detail_sheet.column_dimensions[chr(64 + index)].width = width

        DESKTOP_PATH.mkdir(parents=True, exist_ok=True)
        stem = f"{date.today():%Y-%m-%d}_납기관리 리스트"
        destination = DESKTOP_PATH / f"{stem}.xlsx"
        suffix = 1
        while destination.exists():
            destination = DESKTOP_PATH / f"{stem} ({suffix}).xlsx"
            suffix += 1
        workbook.save(destination)
        self.send_json(200, {"ok": True, "count": len(rows), "filename": destination.name, "path": str(destination)})

    def update_delivery_management(self, body: dict) -> None:
        rows = body.get("rows")
        if not isinstance(rows, list) or not rows or len(rows) > 1000:
            self.send_json(400, {"ok": False, "error": "invalid_rows"})
            return
        now = datetime.now().astimezone().isoformat(timespec="seconds")
        updated = 0
        noted = 0
        unchanged = 0
        dirty = False
        with DELIVERY_LOCK:
            payload = read_delivery_state()
            records = payload.setdefault("records", {})
            for raw in rows:
                item_code = str(raw.get("itemCode") or "").strip()
                spec = str(raw.get("spec") or "").strip()
                adjusted_date = str(raw.get("adjustedDate") or "").strip()
                note = str(raw.get("note") or "").strip()
                if not item_code or (not adjusted_date and not note):
                    continue
                if adjusted_date:
                    try:
                        date.fromisoformat(adjusted_date)
                    except ValueError:
                        self.send_json(400, {"ok": False, "error": "invalid_adjusted_date", "itemCode": item_code})
                        return
                key = delivery_key(item_code, spec)
                record = records.setdefault(key, {
                    "itemCode": item_code,
                    "spec": spec,
                    "confirmedDate": "",
                    "history": [],
                })
                if adjusted_date:
                    previous = str(record.get("confirmedDate") or raw.get("confirmedDate") or "").strip()
                    if previous != adjusted_date:
                        requested = str(raw.get("requestedDate") or "").strip()
                        record.setdefault("history", []).append({
                            "at": now,
                            "type": "date",
                            "text": (
                                f"납기일 수정 {previous} → {adjusted_date}"
                                if previous
                                else f"납기요청일 {requested or '-'} → 납기확정일 {adjusted_date}"
                            ),
                        })
                        updated += 1
                        dirty = True
                    else:
                        unchanged += 1
                    record["confirmedDate"] = adjusted_date
                if note:
                    history = record.setdefault("history", [])
                    if not any(entry.get("type") == "note" and str(entry.get("text") or "").strip() == note for entry in history):
                        history.append({"at": now, "type": "note", "text": note})
                        noted += 1
                        dirty = True
                    else:
                        unchanged += 1
                record["itemCode"] = item_code
                record["spec"] = spec
                record["requestNos"] = list(raw.get("requestNos") or [])
                record["orderNos"] = list(raw.get("orderNos") or [])
                record["waitingStatus"] = str(raw.get("waitingStatus") or "").strip()
            if dirty:
                write_delivery_state(payload)
                write_delivery_confirmations(records)
        self.send_json(200, {"ok": True, **(body.get("importStats") or {}), "updated": updated, "noted": noted, "unchanged": unchanged})

    def append_delivery_note(self, body: dict) -> None:
        item_code = str(body.get("itemCode") or "").strip()
        spec = str(body.get("spec") or "").strip()
        note = str(body.get("note") or "").strip()
        if not item_code or not note or len(note) > 1000:
            self.send_json(400, {"ok": False, "error": "invalid_note"})
            return
        key = delivery_key(item_code, spec)
        with DELIVERY_LOCK:
            payload = read_delivery_state()
            records = payload.setdefault("records", {})
            record = records.setdefault(key, {
                "itemCode": item_code,
                "spec": spec,
                "confirmedDate": "",
                "history": [],
            })
            record.setdefault("history", []).append({
                "at": datetime.now().astimezone().isoformat(timespec="seconds"),
                "type": "note",
                "text": note,
            })
            write_delivery_state(payload)
        self.send_json(200, {"ok": True, "record": record})

    def reconcile_delivery_management(self, body: dict) -> None:
        keys = body.get("keys")
        if not isinstance(keys, list):
            self.send_json(400, {"ok": False, "error": "invalid_keys"})
            return
        removed = 0
        with DELIVERY_LOCK:
            payload = read_delivery_state()
            records = payload.setdefault("records", {})
            for key in keys:
                if records.pop(str(key).upper(), None) is not None:
                    removed += 1
            if removed:
                write_delivery_state(payload)
                write_delivery_confirmations(records)
        self.send_json(200, {"ok": True, "removed": removed})


def main() -> None:
    parser = argparse.ArgumentParser(description="Serve Lidding Foil Planner")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", default=8877, type=int)
    parser.add_argument("--no-open", action="store_true")
    parser.add_argument("--no-monitor", action="store_true")
    args = parser.parse_args()

    monitor_process = None
    if not args.no_monitor:
        creation_flags = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
        monitor_process = subprocess.Popen(
            [sys.executable, str(MONITOR_PATH)],
            cwd=ROOT,
            env=os.environ.copy(),
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=creation_flags,
        )

    os.chdir(WEB_ROOT)
    server = ThreadingHTTPServer((args.host, args.port), DashboardHandler)
    url = f"http://{args.host}:{args.port}/"
    if not args.no_open:
        webbrowser.open_new_tab(url)
    try:
        server.serve_forever()
    finally:
        if monitor_process and monitor_process.poll() is None:
            monitor_process.terminate()


if __name__ == "__main__":
    main()
