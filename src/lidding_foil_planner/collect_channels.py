from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlencode
from urllib.request import Request, urlopen

import psycopg
from openpyxl import load_workbook
from psycopg.types.json import Jsonb

from .config import Settings


APS_PROCESS = "45"
PRODUCTION_PROCESS = "55"
LIDDING_PATTERN = re.compile(r"리드지|lidding|foil", re.IGNORECASE)
HEADER_ROW = 4


def _sha(value: Any) -> str:
    raw = json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str
    ).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _decimal(value: Any) -> Decimal:
    return Decimal("0") if value is None or _text(value) == "" else Decimal(str(value))


def _date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(_text(value)[:10])


def _optional_date(value: Any) -> date | None:
    return None if value is None or _text(value) == "" else _date(value)


def _request_bytes(url: str, timeout: int, accept: str) -> bytes:
    request = Request(
        url,
        headers={"Accept": accept, "User-Agent": "lidding-foil-planner/0.1.0"},
        method="GET",
    )
    with urlopen(request, timeout=timeout) as response:
        if response.status != 200:
            raise RuntimeError(f"HTTP {response.status}: {url}")
        return response.read()


def _request_json(url: str, timeout: int) -> dict[str, Any]:
    return json.loads(_request_bytes(url, timeout, "application/json"))


def _request_workbook(url: str, timeout: int):
    content = _request_bytes(
        url,
        timeout,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return load_workbook(BytesIO(content), read_only=True, data_only=True)


def _sheet_rows(workbook) -> tuple[dict[str, int], Iterable[tuple[Any, ...]]]:
    sheet = workbook[workbook.sheetnames[0]]
    header = next(
        sheet.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True)
    )
    index = {_text(value): position for position, value in enumerate(header) if _text(value)}
    return index, sheet.iter_rows(min_row=HEADER_ROW + 1, values_only=True)


def fetch_aps(settings: Settings) -> dict[str, Any]:
    url = (
        f"{settings.api_base_url}/api/aps-plan?"
        + urlencode({"oper": APS_PROCESS, "limit": 0})
    )
    payload = _request_json(url, settings.http_timeout_seconds)
    rows = payload.get("rows") or []
    if payload.get("truncated") or payload.get("returned_count") != payload.get("total_count"):
        raise RuntimeError("APS hydration response is incomplete")
    if len(rows) != int(payload.get("total_count", -1)):
        raise RuntimeError("APS hydration row count does not match total_count")

    prepared = []
    for row in rows:
        qty = _decimal(row.get("plan_qty"))
        p_code = _text(row.get("item_cd_5"))
        if _text(row.get("oper_id")) != APS_PROCESS or qty <= 0 or not p_code.startswith("P"):
            continue
        identity = {
            "plan_date": row.get("plan_date"),
            "res_id": row.get("res_id"),
            "demand_id": row.get("demand_id"),
            "item_id": row.get("item_id"),
            "due_date": row.get("due_date"),
        }
        prepared.append(
            {
                "row_key": _sha(identity),
                "source_refreshed_at": payload["source_refreshed_at"],
                "plan_date": _date(row["plan_date"]),
                "oper_id": APS_PROCESS,
                "res_id": _text(row.get("res_id")) or None,
                "site_name": _text(row.get("res_site_id")) or None,
                "demand_id": _text(row.get("demand_id")) or None,
                "item_id": _text(row.get("item_id")),
                "p_code": p_code,
                "item_name": _text(row.get("item_name")) or _text(row.get("demand_item_name")) or None,
                "shortage_qty": qty,
                "due_date": _optional_date(row.get("due_date")),
                "demand_type": _text(row.get("demand_type")) or None,
                "sales_order_no": _text(row.get("so_id")) or None,
            }
        )
    return {
        "rows": prepared,
        "source_total_count": payload["total_count"],
        "source_refreshed_at": payload["source_refreshed_at"],
        "p_codes": {row["p_code"] for row in prepared},
        "total_shortage_qty": sum((row["shortage_qty"] for row in prepared), Decimal("0")),
    }


def fetch_production(
    settings: Settings, date_from: date, date_to: date
) -> dict[str, Any]:
    query = urlencode({"date_from": date_from.isoformat(), "date_to": date_to.isoformat()})
    url = f"{settings.api_base_url}/api/production-performance/excel?{query}"
    workbook = _request_workbook(url, settings.http_timeout_seconds)
    index, rows = _sheet_rows(workbook)
    required = ["생산일자", "공정코드", "공장", "품목코드", "지시수량"]
    missing = [name for name in required if name not in index]
    if missing:
        raise RuntimeError(f"Production workbook is missing columns: {missing}")

    aggregates: dict[tuple[date, str, str], dict[str, Any]] = defaultdict(
        lambda: {"qty": Decimal("0"), "count": 0}
    )
    all_rows = filtered_rows = 0
    extracted_at = None
    extracted_index = index.get("원천_추출시각")
    for values in rows:
        if not any(value is not None for value in values):
            continue
        all_rows += 1
        if _text(values[index["공정코드"]]) != PRODUCTION_PROCESS:
            continue
        filtered_rows += 1
        key = (
            _date(values[index["생산일자"]]),
            _text(values[index["공장"]]),
            _text(values[index["품목코드"]]),
        )
        if not key[1] or not key[2]:
            raise RuntimeError(f"Process 55 row has an empty key: {key}")
        aggregates[key]["qty"] += _decimal(values[index["지시수량"]])
        aggregates[key]["count"] += 1
        if extracted_index is not None and _text(values[extracted_index]):
            extracted_at = max(extracted_at or "", _text(values[extracted_index]))

    prepared = []
    for (production_date, factory_code, item_code), value in aggregates.items():
        identity = {
            "production_date": production_date.isoformat(),
            "factory_code": factory_code,
            "item_code": item_code,
        }
        prepared.append(
            {
                "row_key": _sha(identity),
                "production_date": production_date,
                "factory_code": factory_code,
                "item_code": item_code,
                "instruction_qty": value["qty"],
                "source_row_count": value["count"],
                "source_extracted_at": extracted_at,
            }
        )
    if not prepared:
        raise RuntimeError("Production workbook has no process 55 rows")
    return {
        "rows": prepared,
        "source_total_count": all_rows,
        "filtered_source_rows": filtered_rows,
        "source_refreshed_at": extracted_at,
    }


def fetch_bom(settings: Settings, aps_p_codes: set[str]) -> dict[str, Any]:
    url = f"{settings.api_base_url}/api/bom-explosion/excel"
    workbook = _request_workbook(url, settings.http_timeout_seconds)
    index, rows = _sheet_rows(workbook)
    required = [
        "조회기준코드",
        "하위품목코드",
        "하위품목명",
        "규격",
        "소요량",
        "사용여부",
    ]
    missing = [name for name in required if name not in index]
    if missing:
        raise RuntimeError(f"BOM workbook is missing columns: {missing}")

    prepared = []
    all_rows = 0
    extracted_at = None
    extracted_index = index.get("원천_추출시각")
    for values in rows:
        if not any(value is not None for value in values):
            continue
        all_rows += 1
        p_code = _text(values[index["조회기준코드"]])
        bs_code = _text(values[index["하위품목코드"]])
        name = _text(values[index["하위품목명"]])
        use_yn = _text(values[index["사용여부"]])
        if (
            p_code not in aps_p_codes
            or not bs_code.startswith("BS")
            or use_yn != "Y"
            or not LIDDING_PATTERN.search(name)
        ):
            continue
        specification = _text(values[index["규격"]]) or None
        identity = {"p_code": p_code, "bs_code": bs_code, "specification": specification}
        prepared.append(
            {
                "row_key": _sha(identity),
                "p_code": p_code,
                "bs_code": bs_code,
                "lidding_foil_name": name,
                "specification": specification,
                "usage_qty": _decimal(values[index["소요량"]]),
                "use_yn": "Y",
                "source_extracted_at": None,
            }
        )
        if extracted_index is not None and _text(values[extracted_index]):
            extracted_at = max(extracted_at or "", _text(values[extracted_index]))

    for row in prepared:
        row["source_extracted_at"] = extracted_at
    mapped = {row["p_code"] for row in prepared}
    return {
        "rows": prepared,
        "source_total_count": all_rows,
        "source_refreshed_at": extracted_at,
        "mapped_p_codes": mapped,
        "missing_p_codes": sorted(aps_p_codes - mapped),
    }


def _schema_sql() -> str:
    root = Path(__file__).resolve().parents[2]
    return (root / "sql" / "002_channels.sql").read_text(encoding="utf-8")


def _connect(settings: Settings):
    return psycopg.connect(
        host=settings.pg_host,
        port=settings.pg_port,
        dbname=settings.pg_database,
        user=settings.pg_user,
        password=settings.pg_password,
        connect_timeout=settings.pg_connect_timeout,
        application_name="lidding-foil-planner",
    )


def _create_run(
    cursor,
    channel: str,
    date_from: date | None,
    date_to: date | None,
    refreshed_at: str | None,
    source_count: int,
    received_count: int,
    metadata: dict[str, Any],
) -> int:
    cursor.execute(
        """
        insert into lidding_foil_planner.collection_run (
            channel, status, requested_date_from, requested_date_to,
            source_refreshed_at, source_total_count, received_count, metadata
        ) values (%s, 'running', %s, %s, %s, %s, %s, %s)
        returning run_id
        """,
        (channel, date_from, date_to, refreshed_at, source_count, received_count, Jsonb(metadata)),
    )
    return int(cursor.fetchone()[0])


def _replace(cursor, table: str, columns: tuple[str, ...], rows: list[dict[str, Any]], run_id: int) -> int:
    cursor.execute(f"select count(*) from lidding_foil_planner.{table}")
    replaced = int(cursor.fetchone()[0])
    cursor.execute(f"delete from lidding_foil_planner.{table}")
    with cursor.copy(
        f"copy lidding_foil_planner.{table} ({', '.join(columns)}, run_id) from stdin"
    ) as copy:
        for row in rows:
            copy.write_row(tuple(row.get(column) for column in columns) + (run_id,))
    return replaced


def load_bundle(
    settings: Settings,
    aps: dict[str, Any],
    production: dict[str, Any],
    bom: dict[str, Any],
    date_from: date,
    date_to: date,
) -> dict[str, int]:
    aps_columns = (
        "row_key", "source_refreshed_at", "plan_date", "oper_id", "res_id",
        "site_name", "demand_id", "item_id", "p_code", "item_name",
        "shortage_qty", "due_date", "demand_type", "sales_order_no",
    )
    production_columns = (
        "row_key", "production_date", "factory_code", "item_code",
        "instruction_qty", "source_row_count", "source_extracted_at",
    )
    bom_columns = (
        "row_key", "p_code", "bs_code", "lidding_foil_name", "specification",
        "usage_qty", "use_yn", "source_extracted_at",
    )

    with _connect(settings) as connection:
        with connection.cursor() as cursor:
            cursor.execute("set time zone 'Asia/Seoul'")
            cursor.execute(_schema_sql())
        connection.commit()

        with connection.transaction():
            with connection.cursor() as cursor:
                cursor.execute("select pg_advisory_xact_lock(hashtext(%s))", ("lidding-foil-planner:bundle",))
                aps_run = _create_run(
                    cursor, "aps_hydration_shortage", None, None,
                    aps["source_refreshed_at"], aps["source_total_count"], len(aps["rows"]),
                    {"oper_id": APS_PROCESS, "p_code_count": len(aps["p_codes"]), "total_shortage_qty": str(aps["total_shortage_qty"])},
                )
                production_run = _create_run(
                    cursor, "production_performance_55", date_from, date_to,
                    production["source_refreshed_at"], production["source_total_count"], len(production["rows"]),
                    {"oper_id": PRODUCTION_PROCESS, "filtered_source_rows": production["filtered_source_rows"], "grain": "production_date+factory_code+item_code"},
                )
                bom_run = _create_run(
                    cursor, "bom_lidding_foil", None, None,
                    bom["source_refreshed_at"], bom["source_total_count"], len(bom["rows"]),
                    {"scope": "current APS hydration P-codes", "mapped_p_codes": len(bom["mapped_p_codes"]), "missing_p_codes": bom["missing_p_codes"]},
                )

                replaced = {
                    "aps": _replace(cursor, "aps_hydration_shortage", aps_columns, aps["rows"], aps_run),
                    "production": _replace(cursor, "production_performance_55", production_columns, production["rows"], production_run),
                    "bom": _replace(cursor, "bom_lidding_foil", bom_columns, bom["rows"], bom_run),
                }
                for run_id, old_count in (
                    (aps_run, replaced["aps"]),
                    (production_run, replaced["production"]),
                    (bom_run, replaced["bom"]),
                ):
                    cursor.execute(
                        "update lidding_foil_planner.collection_run set status='success', replaced_count=%s, finished_at=now() where run_id=%s",
                        (old_count, run_id),
                    )
    return replaced


def write_status(
    aps: dict[str, Any], production: dict[str, Any], bom: dict[str, Any], date_from: date, date_to: date
) -> None:
    root = Path(__file__).resolve().parents[2]
    path = root / "web" / "data" / "collection-status.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    status = {
        "updatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "aps": {"process": 45, "rows": len(aps["rows"]), "pCodes": len(aps["p_codes"]), "shortageQty": float(aps["total_shortage_qty"]), "sourceRefreshedAt": aps["source_refreshed_at"]},
        "production": {"process": 55, "dateFrom": date_from.isoformat(), "dateTo": date_to.isoformat(), "rows": len(production["rows"]), "sourceRows": production["filtered_source_rows"]},
        "bom": {"rows": len(bom["rows"]), "mappedPCodes": len(bom["mapped_p_codes"]), "missingPCodes": bom["missing_p_codes"]},
    }
    path.write_text(json.dumps(status, ensure_ascii=False, indent=2), encoding="utf-8")


def main(argv: list[str] | None = None) -> int:
    default_to = date.today() - timedelta(days=1)
    parser = argparse.ArgumentParser(description="Collect APS, production 55 and lidding foil BOM channels")
    parser.add_argument("--date-from", type=date.fromisoformat, default=default_to - timedelta(days=6))
    parser.add_argument("--date-to", type=date.fromisoformat, default=default_to)
    args = parser.parse_args(argv)
    settings = Settings.from_env()

    print("Fetching APS hydration shortage...")
    aps = fetch_aps(settings)
    print("Fetching seven-day process 55 production performance...")
    production = fetch_production(settings, args.date_from, args.date_to)
    print("Fetching P-code to BS-code lidding foil BOM...")
    bom = fetch_bom(settings, aps["p_codes"])
    print("Writing validated channel tables to DMZ PostgreSQL...")
    replaced = load_bundle(settings, aps, production, bom, args.date_from, args.date_to)
    write_status(aps, production, bom, args.date_from, args.date_to)
    print(
        json.dumps(
            {
                "aps_rows": len(aps["rows"]),
                "aps_p_codes": len(aps["p_codes"]),
                "production_rows": len(production["rows"]),
                "production_source_rows": production["filtered_source_rows"],
                "bom_rows": len(bom["rows"]),
                "mapped_p_codes": len(bom["mapped_p_codes"]),
                "missing_p_codes": len(bom["missing_p_codes"]),
                "replaced": replaced,
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())

